"""Read Excel File
"""
import openpyxl
# Path
data = openpyxl.load_workbook('E:/tempdata/kldata.xlsx')
# Get Sheet name from the excel file
data.sheetnames
# Get count of rows and columns from each sheet

for i in data.sheetnames:
    print("rows in", i , data[i].max_row)
    print("columns in", i, data[i].max_column)
